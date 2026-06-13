"""Compare & Match — reconcile extracted trades against a golden source.

This is the capability the BRD puts at the centre of the engagement (~80% of SSG
Ops tasks). After extraction, each trade is:

  1. **matched** by ``trade_id`` against a trusted *golden source* (here a mock
     master blotter; in production GLOSS / OBI / FO systems);
  2. **enriched** — any field the email/attachment did not carry is populated
     from the golden record;
  3. **compared** — the fields the extraction *did* carry are checked against the
     golden values (agree / disagree), producing field-level diffs and a
     confidence score;
  4. **classified** into a status: MATCHED · ENRICHED · NEAR_MATCH · BREAK ·
     UNMATCHED.

Design contract:
* **Never raises** — a missing golden file or one bad row yields an empty source
  / an UNMATCHED result, never an exception.
* **Source-agnostic** — the golden rows are normalized with the *same* coercers
  as the extractor, so values compare apples-to-apples regardless of origin.
"""

from dataclasses import dataclass, field
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, Optional

from agent.attachment_extractor import (
    _DATE_FIELDS,
    _NUMERIC_FIELDS,
    _coerce_date,
    _normalize_row,
)

# Match statuses (string constants so they serialize cleanly to JSON / the UI).
MATCHED = "MATCHED"        # golden found, every compared field agrees, nothing to fill
ENRICHED = "ENRICHED"      # golden found, compared fields agree, ≥1 missing field filled
NEAR_MATCH = "NEAR_MATCH"  # golden found, some compared fields disagree (confidence high-ish)
BREAK = "BREAK"            # golden found, compared fields disagree badly (confidence low)
UNMATCHED = "UNMATCHED"    # no golden record for this trade id


def _blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def _norm_str(v: Any) -> str:
    return " ".join(str(v).strip().lower().split())


def values_agree(field_name: str, ev: Any, gv: Any, *, tolerance: float) -> bool:
    """Type-aware equality: numbers within tolerance, dates canonicalized, strings folded."""
    if field_name in _NUMERIC_FIELDS:
        try:
            a, b = float(str(ev).replace(",", "")), float(str(gv).replace(",", ""))
        except (TypeError, ValueError):
            return _norm_str(ev) == _norm_str(gv)
        if b == 0:
            return abs(a - b) <= tolerance
        return abs(a - b) / abs(b) <= tolerance
    if field_name in _DATE_FIELDS:
        return _coerce_date(ev) == _coerce_date(gv)
    return _norm_str(ev) == _norm_str(gv)


# --------------------------------------------------------------------------
# Golden source loader (cached by path + mtime so the workbook is read once)
# --------------------------------------------------------------------------

@lru_cache(maxsize=8)
def _load_golden(path: str, sheet: str, key: str, _mtime: float) -> Dict[str, Dict[str, Any]]:
    """Read the golden workbook into ``{trade_id: normalized_row}``. Never raises."""
    import openpyxl  # lazy import (only Compare & Match needs it)

    index: Dict[str, Dict[str, Any]] = {}
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:  # noqa: BLE001 - missing/corrupt golden file => empty source
        return index
    try:
        ws = wb[sheet] if sheet in wb.sheetnames else wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return index

    headers = [str(h or "") for h in rows[0]]
    for cells in rows[1:]:
        raw = {headers[i]: (cells[i] if i < len(cells) else None) for i in range(len(headers))}
        norm = _normalize_row(raw)
        tid = str(norm.get(key) or "").strip()
        if tid:
            index[tid] = norm
    return index


class GoldenSource:
    """Trusted master dataset, indexed by trade id. Stands in for GLOSS/OBI/FO."""

    def __init__(self, path: str, *, sheet: str = "FX Options Trade Blotter", key: str = "trade_id"):
        self.path = path
        self.sheet = sheet
        self.key = key

    @property
    def available(self) -> bool:
        return Path(self.path).exists()

    def _index(self) -> Dict[str, Dict[str, Any]]:
        p = Path(self.path)
        mtime = p.stat().st_mtime if p.exists() else 0.0
        return _load_golden(str(self.path), self.sheet, self.key, mtime)

    def __len__(self) -> int:
        return len(self._index())

    def lookup(self, trade_id: str) -> Optional[Dict[str, Any]]:
        if not trade_id:
            return None
        return self._index().get(str(trade_id).strip())


# --------------------------------------------------------------------------
# Result + matcher
# --------------------------------------------------------------------------

@dataclass
class FieldDiff:
    field: str
    extracted: Any
    golden: Any
    agree: bool

    def as_dict(self) -> Dict[str, Any]:
        return {"field": self.field, "extracted": self.extracted,
                "golden": self.golden, "agree": self.agree}


@dataclass
class MatchResult:
    trade_id: str
    status: str
    confidence: float
    golden_found: bool
    completed: Dict[str, Any] = field(default_factory=dict)   # extracted ∪ golden-filled
    filled_fields: List[str] = field(default_factory=list)     # fields populated from golden
    diffs: List[FieldDiff] = field(default_factory=list)       # compared (present-in-both) fields
    compared_fields: int = 0
    agreed_fields: int = 0

    def as_dict(self) -> Dict[str, Any]:
        return {
            "trade_id": self.trade_id,
            "status": self.status,
            "confidence": self.confidence,
            "golden_found": self.golden_found,
            "filled_fields": self.filled_fields,
            "filled_count": len(self.filled_fields),
            "compared_fields": self.compared_fields,
            "agreed_fields": self.agreed_fields,
            "diffs": [d.as_dict() for d in self.diffs],
            "breaks": [d.as_dict() for d in self.diffs if not d.agree],
            "completed": self.completed,
        }


class CompareMatcher:
    """Match → enrich → compare a list of extracted trades against a GoldenSource."""

    def __init__(self, golden: GoldenSource, *, key: str = "trade_id",
                 match_fields: Optional[List[str]] = None,
                 tolerance: float = 0.01, break_threshold: float = 0.6,
                 enrich: bool = True):
        self.golden = golden
        self.key = key
        self.match_fields = list(match_fields) if match_fields else []
        self.tolerance = tolerance
        self.break_threshold = break_threshold
        self.enrich = enrich

    def match_one(self, extracted: Dict[str, Any]) -> MatchResult:
        tid = str(extracted.get(self.key) or "").strip()
        golden = self.golden.lookup(tid)
        completed = dict(extracted)

        if golden is None:
            return MatchResult(trade_id=tid, status=UNMATCHED, confidence=0.0,
                               golden_found=False, completed=completed)

        filled: List[str] = []
        diffs: List[FieldDiff] = []

        for fname, gv in golden.items():
            if fname == self.key:
                continue
            ev = extracted.get(fname)
            if _blank(ev):
                # Enrichment: the email/attachment didn't carry this field.
                if self.enrich and not _blank(gv):
                    completed[fname] = gv
                    filled.append(fname)
            elif fname in self.match_fields and not _blank(gv):
                # Comparison: both sides carry it → check agreement.
                diffs.append(FieldDiff(fname, ev, gv,
                                       values_agree(fname, ev, gv, tolerance=self.tolerance)))

        compared = len(diffs)
        agreed = sum(1 for d in diffs if d.agree)
        confidence = (agreed / compared) if compared else 1.0

        if compared and agreed < compared:
            status = BREAK if confidence < self.break_threshold else NEAR_MATCH
        elif filled:
            status = ENRICHED
        else:
            status = MATCHED

        return MatchResult(
            trade_id=tid, status=status, confidence=round(confidence, 3),
            golden_found=True, completed=completed, filled_fields=sorted(filled),
            diffs=diffs, compared_fields=compared, agreed_fields=agreed,
        )

    def match(self, trades: List[Dict[str, Any]]) -> List[MatchResult]:
        return [self.match_one(t) for t in trades]

    @staticmethod
    def summarize(results: List[MatchResult]) -> Dict[str, Any]:
        by_status: Dict[str, int] = {}
        for r in results:
            by_status[r.status] = by_status.get(r.status, 0) + 1
        total = len(results)
        avg_conf = round(sum(r.confidence for r in results) / total, 3) if total else 0.0
        return {
            "total": total,
            "by_status": by_status,
            "matched_or_enriched": by_status.get(MATCHED, 0) + by_status.get(ENRICHED, 0),
            "breaks": by_status.get(BREAK, 0) + by_status.get(NEAR_MATCH, 0),
            "unmatched": by_status.get(UNMATCHED, 0),
            "fields_filled": sum(len(r.filled_fields) for r in results),
            "avg_confidence": avg_conf,
        }
