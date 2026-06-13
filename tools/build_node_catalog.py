"""Generate the Workflow-Builder Node Catalog (Excel).

A design artifact for the "agent marketplace / composable workflow builder" vision.
It enumerates every block (node), its inputs/outputs (ports), configuration, the
data contracts that flow on the edges, the valid connection permutations, example
assembled workflows, and how each maps to what's BUILT today vs. the BRD's future
scope.

Run:  python tools/build_node_catalog.py
Out:  docs/Workflow_Builder_Node_Catalog.xlsx
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = Path(__file__).resolve().parents[1] / "docs" / "Workflow_Builder_Node_Catalog.xlsx"

# ───────────────────────── palette ─────────────────────────
NAVY   = "1A3260"   # header
NAVY_D = "112244"
RED    = "C8102E"   # Nomura accent
WHITE  = "FFFFFF"
GREY_HDR = "33415A"

STATUS_FILL = {
    "Built":   "C6EFCE",   # green
    "Partial": "FFEB9C",   # amber
    "Planned": "BDD7EE",   # blue
    "Future":  "E7E0EC",   # lilac/grey
}
STATUS_FONT = {
    "Built":   "1E6B34",
    "Partial": "9C6500",
    "Planned": "1F4E79",
    "Future":  "5B4B7A",
}

CAT_FILL = {
    "Source / Trigger":          "DDEBF7",
    "Ingest / Parse":            "E2EFDA",
    "Classify / Reduce":         "FCE4D6",
    "Extract / Transform":       "FFF2CC",
    "Decision / Routing":        "FBE5D6",
    "Compare & Match":           "D9E1F2",
    "Human-in-the-Loop":         "EAD1DC",
    "Action / Output":           "D6E4F0",
    "Integration":              "E2E2F0",
    "Memory / Learning":         "DDF0EE",
    "Governance / Observability":"F2F2F2",
    "Utility / Flow":            "EDEDED",
}

thin = Side(style="thin", color="C9CED6")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)


# ───────────────────────── data: data contracts ─────────────────────────
# canonical objects that flow on the edges between nodes
CONTRACTS = [
    ("TRIGGER", "A run signal that starts a workflow (no business payload).",
     "run_id, triggered_by, timestamp, params{}",
     "Triggers (Manual, Scheduler, ServiceNow)", "Any ingest / first node"),
    ("RAW_MESSAGE", "A normalized inbound message (email / SWIFT / API record).",
     "message_id, subject, sender, received_at, body, html_body, attachments[], source_file",
     "Email/SWIFT/REST ingestors", "Classifier, Field Extractor, Dedup"),
    ("DOCUMENT", "A binary attachment / file with metadata.",
     "filename, bytes, mime_type, size_bytes, source_message_id",
     "File ingestor, message attachments", "Attachment Parser, PDF/OCR, Doc Parser"),
    ("CLASSIFIED_ITEM", "A message plus its classification verdict.",
     "…RAW_MESSAGE, label(RELEVANT/AMBIGUOUS/IRRELEVANT), confidence, reason, matched_keywords[], trade_id",
     "Classifier, Intent Classifier", "Decider, Extractor, HITL, Dedup"),
    ("TRADE_RECORD", "A structured trade row (the extraction output).",
     "trade_id, uti, currency_pair, buy_sell, notional_amount, notional_ccy, counterparty, value_date/settlement_date, rate, asset, …",
     "Extractor, Attachment Parser, Normalizer", "Compare & Match, Writer, Manifest"),
    ("GOLDEN_RECORD", "A trusted record fetched from a golden source.",
     "source(GLOSS/OBI/FO), trade_id, fields{}, retrieved_at",
     "Golden Source Query nodes", "Compare & Match, SSI Cross-Validator"),
    ("MATCH_RESULT", "Outcome of comparing extracted vs golden data.",
     "trade_id, status(matched/near-match/break), confidence, field_diffs[], reason",
     "Compare & Match, SSI Cross-Validator", "Confidence Router, Exception Classifier, HITL"),
    ("REVIEW_TASK", "A unit of work pushed to a human review queue.",
     "case_id, payload{}, assignee/role, status(awaiting/…), sla",
     "Man in the Loop, Approval Gate", "(human review UI)"),
    ("REVIEW_DECISION", "A human's adjudication of a review task.",
     "case_id, action(approve/reject/override), comments, actor, role, timestamp",
     "HITL nodes (after human acts)", "Feedback Loop, Response Drafter, ServiceNow"),
    ("NOTIFICATION", "An outbound message to a person or system.",
     "channel(email/teams), to, subject, body, attachments[]",
     "Response Drafter, Alert Sender, EMODEST", "Email Sender, ServiceNow"),
    ("CASE_MANIFEST", "The Phase-2 handoff contract for one case.",
     "trade_id, asset_class, paths{}, attachments[], classification{}, extraction{}, ready_for_extraction",
     "Case / Manifest Builder", "Writer, downstream extraction agent"),
    ("AUDIT_EVENT", "An append-only compliance event.",
     "event_id, action, actor, resource, outcome, correlation_id, timestamp, details{}",
     "Audit Logger, most action nodes", "WORM/SIEM, Observability"),
    ("REPORT", "A generated report / MIS output.",
     "title, rows[], columns[], format(xlsx/pdf/json), generated_at",
     "Report Generator, Evaluator", "Email Sender, Writer"),
    ("DATASET", "Generic tabular rows (DB query / spreadsheet / API result).",
     "rows[], schema{}, source",
     "REST/TRAM/Snowflake ingestors, Writer", "Normalizer, Extractor, Report"),
    ("ANY", "Wildcard — passes through whatever it receives.",
     "(pass-through)",
     "Utility / flow nodes", "any node"),
]
CONTRACT_TYPES = [c[0] for c in CONTRACTS]

# explicit cross-type compatibility (besides exact match and ANY⇄*)
COMPAT_EXTRA = {
    ("RAW_MESSAGE", "DATASET"),
    ("CLASSIFIED_ITEM", "RAW_MESSAGE"),
    ("CLASSIFIED_ITEM", "TRADE_RECORD"),
    ("DOCUMENT", "DATASET"),
    ("TRADE_RECORD", "DATASET"),
    ("GOLDEN_RECORD", "DATASET"),
    ("MATCH_RESULT", "DATASET"),
    ("CASE_MANIFEST", "DATASET"),
    ("REPORT", "DATASET"),
    ("REVIEW_DECISION", "DATASET"),
}


def compatible(out_t, in_t):
    if out_t == in_t:
        return "EXACT"
    if out_t == "ANY" or in_t == "ANY":
        return "ANY"
    if (out_t, in_t) in COMPAT_EXTRA:
        return "ADAPT"
    return ""


# ───────────────────────── data: nodes ─────────────────────────
# (id, name, category, theme, status, inputs[(port,type,req)], outputs[(port,type)],
#  config[(param,type,options,default,required,notes)], code_map, ref, downstream)
N = []


def node(id, name, cat, theme, status, ins, outs, cfg, code, ref, down):
    N.append(dict(id=id, name=name, cat=cat, theme=theme, status=status,
                  ins=ins, outs=outs, cfg=cfg, code=code, ref=ref, down=down))


# —— Source / Trigger ——
node("TRG-01", "Manual / Webhook Trigger", "Source / Trigger", "Orchestration", "Future",
     [], [("out", "TRIGGER")],
     [("trigger_type", "enum", "manual | webhook", "manual", "Y", "Starts a run on demand or via HTTP call."),
      ("webhook_secret", "secret", "", "", "N", "Shared secret for webhook auth.")],
     "—", "Marketplace / orchestration", "Any ingest node")

node("TRG-02", "Scheduler (Cron) Trigger", "Source / Trigger", "Orchestration", "Planned",
     [], [("out", "TRIGGER")],
     [("frequency_hours", "int", "", "24", "Y", "Cadence; BRD: 4 cycles / 3 hrs."),
      ("cron", "string", "cron expr", "", "N", "Advanced schedule override.")],
     "config.sync_frequency_hours (captured, not wired)", "BRD §6 / MoM volumes", "Any ingest node")

node("TRG-03", "ServiceNow (NEWS) Trigger", "Source / Trigger", "Orchestration", "Future",
     [], [("out", "TRIGGER"), ("case", "DATASET")],
     [("instance_url", "url", "", "", "Y", "NEWS instance."),
      ("table", "string", "", "incident", "Y", "Watched table / queue."),
      ("event", "enum", "created | updated", "created", "Y", "Fires on case event.")],
     "—", "BRD §3 ServiceNow integration", "Ingest / Classify")

# —— Ingest / Parse ——
node("ING-01", "Email Ingestor — MS Graph API", "Ingest / Parse", "Compare & Match", "Built",
     [("trigger", "TRIGGER", "N")], [("messages", "RAW_MESSAGE")],
     [("source", "enum", "REST API | Microsoft Graph API", "Microsoft Graph API", "Y", "Connector type."),
      ("endpoint_url", "url", "", "graph.microsoft.com/v1.0/users/{mailbox}/messages", "Y", "Mailbox endpoint."),
      ("auth_key", "secret", "OAuth2 client-credentials", "", "Y", "Tenant/client/secret in prod."),
      ("payload", "json", "$filter/$select/$top", "{\"$top\":10}", "N", "OData query."),
      ("store_to", "enum", "Database | File System", "File System", "Y", "Where raw lands.")],
     "connectors/graph_connector.py", "Built", "Classify / Reduce")

node("ING-02", "Email Ingestor — Local (.eml)", "Ingest / Parse", "Compare & Match", "Built",
     [("trigger", "TRIGGER", "N")], [("messages", "RAW_MESSAGE")],
     [("inbox_path", "path", "", "data/raw_emails/inbox", "Y", "Folder of .eml files."),
      ("store_to", "enum", "Database | File System", "File System", "N", "")],
     "connectors/local_connector.py", "Built", "Classify / Reduce")

node("ING-03", "Email Ingestor — IMAP", "Ingest / Parse", "Compare & Match", "Planned",
     [("trigger", "TRIGGER", "N")], [("messages", "RAW_MESSAGE")],
     [("host", "string", "", "", "Y", "IMAP server."),
      ("mailbox", "string", "", "INBOX", "Y", ""),
      ("auth_key", "secret", "", "", "Y", "")],
     "—", "BRD §6.4 IMAP / MS Graph", "Classify / Reduce")

node("ING-04", "REST API Ingestor", "Ingest / Parse", "Reporting & MIS", "Partial",
     [("trigger", "TRIGGER", "N")], [("data", "DATASET")],
     [("endpoint_url", "url", "", "", "Y", "Any REST source."),
      ("method", "enum", "GET | POST", "GET", "Y", ""),
      ("auth_key", "secret", "", "", "N", ""),
      ("payload", "json", "", "{}", "N", "")],
     "UI source option (REST API)", "BRD §6.4 REST APIs", "Normalizer / Extractor")

node("ING-05", "File / Folder Ingestor", "Ingest / Parse", "Compare & Match", "Partial",
     [("trigger", "TRIGGER", "N")], [("documents", "DOCUMENT")],
     [("path", "path", "", "", "Y", "File or folder."),
      ("formats", "csv", ".pdf,.xlsx,.csv,.docx,.eml,.txt", ".xlsx,.csv", "Y", "Allowed types.")],
     "file ingestion (partial)", "BRD §6.4 file ingestion", "Attachment Parser / PDF-OCR")

node("ING-06", "SWIFT Message Ingestor", "Ingest / Parse", "Compare & Match", "Future",
     [("trigger", "TRIGGER", "N")], [("messages", "RAW_MESSAGE"), ("data", "DATASET")],
     [("message_types", "csv", "MT103,MT202,MT202COV", "MT103,MT202", "Y", "SWIFT MT types."),
      ("feed", "string", "", "", "Y", "Feed / queue name.")],
     "—", "BRD §6 SWIFT", "Normalizer / Extractor")

node("ING-07", "CSD / ICSD Report Ingestor", "Ingest / Parse", "Compare & Match", "Future",
     [("trigger", "TRIGGER", "N")], [("documents", "DOCUMENT"), ("data", "DATASET")],
     [("source", "string", "", "", "Y", "Custodian / depository."),
      ("format", "enum", "pdf | csv | xml", "csv", "Y", "")],
     "—", "BRD §6 CSD/ICSD reports", "PDF-OCR / Normalizer")

node("ING-08", "TRAM Feed Ingestor", "Ingest / Parse", "Compare & Match", "Future",
     [("trigger", "TRIGGER", "N")], [("data", "DATASET")],
     [("feed", "string", "", "", "Y", "TRAM feed."),
      ("since", "string", "", "", "N", "Incremental cursor.")],
     "—", "BRD §6 TRAM", "Normalizer / Compare & Match")

node("ING-09", "Snowflake Query Ingestor", "Ingest / Parse", "Reporting & MIS", "Future",
     [("trigger", "TRIGGER", "N")], [("data", "DATASET")],
     [("query", "sql", "", "", "Y", "SQL against Snowflake."),
      ("warehouse", "string", "", "", "Y", "")],
     "—", "BRD §6 Snowflake", "Normalizer / Report")

node("ING-10", "Attachment Parser (xlsx / csv)", "Ingest / Parse", "Compare & Match", "Built",
     [("documents", "DOCUMENT", "Y")], [("trades", "TRADE_RECORD")],
     [("supported_exts", "csv", ".xlsx,.xlsm,.csv", ".xlsx,.xlsm,.csv", "Y", "Blotter formats."),
      ("max_rows", "int", "", "10000", "N", "Row cap per file.")],
     "agent/attachment_extractor.py", "Built", "Dedup / Compare & Match / Writer")

node("ING-11", "PDF / OCR Extractor", "Ingest / Parse", "Compare & Match", "Future",
     [("documents", "DOCUMENT", "Y")], [("trades", "TRADE_RECORD"), ("data", "DATASET")],
     [("ocr", "bool", "", "true", "Y", "Render-to-image + OCR (Tesseract)."),
      ("layout", "enum", "text | tables | both", "both", "N", "")],
     "—", "BRD/roadmap: SSI PDFs (vector, need OCR)", "Normalizer / SSI Cross-Validator")

node("ING-12", "Document Parser (docx / txt)", "Ingest / Parse", "Compare & Match", "Future",
     [("documents", "DOCUMENT", "Y")], [("data", "DATASET")],
     [("formats", "csv", ".docx,.txt", ".docx,.txt", "Y", "")],
     "—", "Roadmap", "Normalizer / Extractor")

# —— Classify / Reduce ——
node("CLS-01", "Data Reducer / Classifier (NLP keyword)", "Classify / Reduce", "Compare & Match", "Built",
     [("messages", "RAW_MESSAGE", "Y")], [("classified", "CLASSIFIED_ITEM")],
     [("mode", "enum", "NLP (keyword) | AI | NLP + AI", "NLP (keyword)", "Y", "Scoring engine."),
      ("asset_keywords", "csv", "", "fx trade settlement,deal reference,…", "Y", "+0.5 each."),
      ("subject_keywords", "csv", "", "settlement,confirm,trade,…", "N", "+0.3 (subject)."),
      ("negative_keywords", "csv", "", "birthday,it support,…", "N", "Hard-negative early exit."),
      ("relevant_threshold", "float", "0–1", "0.7", "Y", "≥ → RELEVANT."),
      ("ambiguous_threshold", "float", "0–1", "0.3", "Y", "band → AMBIGUOUS."),
      ("store_back", "enum", "SQLite | Excel", "SQLite", "N", "")],
     "classifier/rule_classifier.py + config/config_store.py", "Built", "Decider / Extractor / HITL")

node("CLS-02", "Intent Classifier (AI / LLM)", "Classify / Reduce", "Compare & Match", "Planned",
     [("messages", "RAW_MESSAGE", "Y")], [("classified", "CLASSIFIED_ITEM")],
     [("model", "enum", "Anthropic Chinou", "claude", "Y", "Context-aware, not keyword."),
      ("labels", "csv", "", "Allege,DK,SSI,Confirmation,Noise", "Y", "Intent set."),
      ("prompt", "text", "", "", "N", "System prompt / few-shot.")],
     "—", "BRD §6.2 context-aware; demote rules to backbone", "Decider / Extractor / HITL")

node("CLS-03", "Deduplicator", "Classify / Reduce", "Compare & Match", "Built",
     [("items", "CLASSIFIED_ITEM", "Y")], [("unique", "CLASSIFIED_ITEM"), ("duplicates", "CLASSIFIED_ITEM")],
     [("keys", "csv", "message_id,trade_id", "message_id,trade_id", "Y", "Idempotency guards."),
      ("store", "path", "", "data/email_index.db", "N", "SQLite index.")],
     "storage/db_index.py", "Built", "Writer / Compare & Match")

node("CLS-04", "Normalizer / Schema Mapper", "Classify / Reduce", "Compare & Match", "Built",
     [("data", "DATASET", "Y")], [("trades", "TRADE_RECORD")],
     [("header_map", "json", "", "{cpty:counterparty,…}", "Y", "Column aliases → canonical."),
      ("date_format", "string", "", "%d-%b-%Y", "N", "Source-agnostic dates."),
      ("number_locale", "string", "", "en", "N", "Comma/locale handling.")],
     "agent/attachment_extractor.py (_normalize_row)", "Built", "Compare & Match / Writer")

node("CLS-05", "Aggregator / Blotter Exploder", "Classify / Reduce", "Compare & Match", "Built",
     [("trades", "TRADE_RECORD", "Y")], [("rows", "TRADE_RECORD")],
     [("mode", "enum", "explode | group | sum", "explode", "Y", "One blotter → N trades."),
      ("group_by", "csv", "", "", "N", "")],
     "extract loop (one row per trade)", "Built", "Writer / Report")

# —— Extract / Transform ——
node("EXT-01", "Field Extractor (Regex / NLP / AI)", "Extract / Transform", "Compare & Match", "Partial",
     [("source", "CLASSIFIED_ITEM", "Y"), ("documents", "DOCUMENT", "N")], [("trades", "TRADE_RECORD")],
     [("input_source", "path", "", "./data/trades.db", "Y", "File/DB or upstream port."),
      ("file_formats", "csv", ".eml,.txt,.xlsx,.csv", ".eml,.txt", "N", ""),
      ("fields", "csv", "", "trade_id,currency_pair,notional_amount,settlement_date,counterparty,trader_name", "Y", "6 BRD fields."),
      ("method", "enum", "NLP Pattern Matching | Regex Rules | AI Extraction", "Regex Rules", "Y", "Extraction engine.")],
     "api/routers/extract.py + client-side parse (body) / attachment_extractor (xlsx)", "Partial", "Compare & Match / Writer / Manifest")

node("EXT-02", "Unstructured → Structured Extractor (LLM)", "Extract / Transform", "Compare & Match", "Future",
     [("any", "ANY", "Y")], [("trades", "TRADE_RECORD"), ("data", "DATASET")],
     [("model", "enum", "Anthropic Chinou", "claude", "Y", "Any format → JSON."),
      ("schema", "json", "", "", "Y", "Target field schema."),
      ("on_missing", "enum", "blank | flag-HITL", "flag-HITL", "N", "BRD: record missing fields for review.")],
     "—", "BRD §6 extraction agent (if required)", "Compare & Match / HITL")

# —— Decision / Routing ——
node("DEC-01", "Decider (field + operator branch)", "Decision / Routing", "Orchestration", "Partial",
     [("in", "ANY", "Y")], [("true", "ANY"), ("false", "ANY")],
     [("field_reference", "string", "", "confidence_score", "Y", "Field from a prior node."),
      ("operator", "enum", "> | >= | < | <= | == | !=", ">", "Y", ""),
      ("value", "string", "", "80", "Y", "Comparison value.")],
     "UI Decider (logic, runtime pending)", "Built (UI) / Partial", "two downstream branches")

node("DEC-02", "Confidence Router", "Decision / Routing", "Orchestration", "Partial",
     [("in", "CLASSIFIED_ITEM", "Y")], [("high", "ANY"), ("low", "ANY")],
     [("threshold", "float", "0–1 or %", "0.9", "Y", "BRD: ≥90% auto-draft; <90% → MO."),
      ("score_field", "string", "", "confidence", "Y", "")],
     "agent confidence routing (partial)", "BRD §3 routing", "Auto-draft / HITL / EMODEST")

node("DEC-03", "Switch / Multi-branch Router", "Decision / Routing", "Orchestration", "Future",
     [("in", "ANY", "Y")], [("branch_n", "ANY")],
     [("field_reference", "string", "", "label", "Y", ""),
      ("cases", "json", "", "{RELEVANT:..,AMBIGUOUS:..}", "Y", "Value → branch map.")],
     "—", "Roadmap", "N downstream branches")

node("DEC-04", "Confidence Scorer", "Decision / Routing", "Compare & Match", "Future",
     [("matches", "MATCH_RESULT", "Y")], [("scored", "MATCH_RESULT")],
     [("weights", "json", "", "{field:weight}", "N", "Per-field weighting."),
      ("scale", "enum", "0-1 | 0-100", "0-100", "N", "")],
     "—", "BRD §3 confidence scoring", "Confidence Router / HITL")

# —— Compare & Match ——
node("CMP-01", "Golden Source Query — GLOSS API", "Compare & Match", "Compare & Match", "Future",
     [("trades", "TRADE_RECORD", "Y")], [("golden", "GOLDEN_RECORD")],
     [("endpoint_url", "url", "", "", "Y", "GLOSS / GLOSS Cash."),
      ("auth_key", "secret", "", "", "Y", ""),
      ("match_keys", "csv", "", "trade_id,counterparty", "Y", "Lookup keys.")],
     "—", "BRD §3 golden source (GLOSS)", "Compare & Match Engine")

node("CMP-02", "Golden Source Query — OBI", "Compare & Match", "Compare & Match", "Future",
     [("trades", "TRADE_RECORD", "Y")], [("golden", "GOLDEN_RECORD")],
     [("endpoint_url", "url", "", "", "Y", "OBI system."),
      ("auth_key", "secret", "", "", "Y", "")],
     "—", "BRD §3 golden source (OBI)", "Compare & Match Engine")

node("CMP-03", "Golden Source Query — FO Systems", "Compare & Match", "Compare & Match", "Future",
     [("trades", "TRADE_RECORD", "Y")], [("golden", "GOLDEN_RECORD")],
     [("system", "string", "", "", "Y", "Front-office source."),
      ("auth_key", "secret", "", "", "Y", "")],
     "—", "BRD §3 golden source (FO)", "Compare & Match Engine")

node("CMP-04", "Compare & Match Engine", "Compare & Match", "Compare & Match", "Future",
     [("extracted", "TRADE_RECORD", "Y"), ("golden", "GOLDEN_RECORD", "Y")], [("results", "MATCH_RESULT")],
     [("fields", "csv", "", "asset,counterparty,notional,value_date,rate", "Y", "6 fields to match."),
      ("tolerance", "json", "", "{notional:0.01}", "N", "Per-field tolerance."),
      ("fuzzy", "bool", "", "true", "N", "Fuzzy + semantic match."),
      ("method", "enum", "field | fuzzy | semantic(AI)", "fuzzy", "Y", "")],
     "—", "BRD §3 Compare & Match — the engagement core", "Confidence Router / Exception / HITL")

node("CMP-05", "SSI Cross-Validator", "Compare & Match", "Compare & Match", "Future",
     [("instructions", "TRADE_RECORD", "Y")], [("results", "MATCH_RESULT")],
     [("golden", "enum", "GLOSS Cash | internal", "GLOSS Cash", "Y", "Receiving-party expectation."),
      ("checks", "csv", "", "swift,account,beneficiary", "Y", "SSI fields to cross-check.")],
     "—", "BRD work driver: SSI Verification", "Exception Classifier / HITL")

node("CMP-06", "Exception / Break Classifier", "Compare & Match", "Alerts & Escalation", "Future",
     [("results", "MATCH_RESULT", "Y")], [("breaks", "MATCH_RESULT")],
     [("rules", "json", "", "", "N", "Break reason taxonomy."),
      ("categories", "csv", "", "matched,near-match,break,missing-data", "Y", "")],
     "—", "BRD: exception flagging / break reasons", "HITL / Alert Sender")

# —— Human-in-the-Loop ——
node("HIL-01", "Man in the Loop (HITL Review)", "Human-in-the-Loop", "Orchestration", "Partial",
     [("in", "ANY", "Y")], [("decision", "REVIEW_DECISION")],
     [("send_email_to", "email", "", "trade-ops@company.com", "Y", "Notify reviewer."),
      ("subject", "string", "", "Low-confidence trade email requires review", "Y", ""),
      ("email_content", "text", "", "Please review and re-classify.", "N", ""),
      ("pause", "bool", "", "true", "Y", "Workflow pauses until a human acts.")],
     "UI Man-in-the-Loop / DecisionDrawer (review UI partial)", "BRD: HITL mandatory gate", "Feedback / Response / ServiceNow")

node("HIL-02", "Approval Gate (RBAC)", "Human-in-the-Loop", "Orchestration", "Future",
     [("in", "ANY", "Y")], [("decision", "REVIEW_DECISION")],
     [("roles", "csv", "", "operator,reviewer,approver", "Y", "Role-based access."),
      ("actions", "csv", "", "approve,reject,override", "Y", ""),
      ("audit", "bool", "", "true", "Y", "Log actor/timestamp/reason.")],
     "—", "BRD FR.04 RBAC + audit", "Feedback / Action / ServiceNow")

node("HIL-03", "EMODEST Broadcast (MO feedback)", "Human-in-the-Loop", "Alerts & Escalation", "Future",
     [("in", "MATCH_RESULT", "Y")], [("notification", "NOTIFICATION"), ("decision", "REVIEW_DECISION")],
     [("distribution", "csv", "", "middle-office", "Y", "MO broadcast list."),
      ("coded_subject", "string", "", "[EMODEST]", "Y", "Coded subject for feedback routing.")],
     "—", "BRD: <90% → EMODEST broadcast to MO", "Feedback Loop")

# —— Action / Output ——
node("ACT-01", "Response Drafter (auto-draft)", "Action / Output", "Optimal Refinement", "Future",
     [("in", "MATCH_RESULT", "Y")], [("draft", "NOTIFICATION")],
     [("model", "enum", "Anthropic Chinou", "claude", "Y", "Generates counterparty reply."),
      ("template", "text", "", "", "N", ""),
      ("require_approval", "bool", "", "true", "N", "BRD: ≥90% auto-draft (still HITL-gated).")],
     "—", "BRD: auto-draft response", "Email Sender / HITL")

node("ACT-02", "Email / Notification Sender", "Action / Output", "Alerts & Escalation", "Partial",
     [("notification", "NOTIFICATION", "Y")], [("audit", "AUDIT_EVENT")],
     [("channel", "enum", "email | teams", "email", "Y", ""),
      ("smtp_or_graph", "string", "", "", "N", "Transport config.")],
     "HITL email (partial)", "Roadmap", "Audit / end")

node("ACT-03", "Alert / Escalation Sender", "Action / Output", "Alerts & Escalation", "Future",
     [("in", "ANY", "Y")], [("notification", "NOTIFICATION")],
     [("severity", "enum", "info | warn | breach", "warn", "Y", ""),
      ("escalation_path", "csv", "", "L1,L2", "N", "")],
     "—", "Capability theme: Alerts & Escalation", "Email Sender / ServiceNow")

node("ACT-04", "Data Writer (SQLite / Excel / DB)", "Action / Output", "Reporting & MIS", "Built",
     [("rows", "TRADE_RECORD", "Y")], [("dataset", "DATASET")],
     [("store", "enum", "SQLite | Excel | Database", "SQLite", "Y", ""),
      ("path", "path", "", "./data/trades.db", "Y", "")],
     "storage/file_store.py + db_index.py", "Built", "Report / end")

node("ACT-05", "Case / Manifest Builder", "Action / Output", "Compare & Match", "Built",
     [("classified", "CLASSIFIED_ITEM", "Y"), ("trades", "TRADE_RECORD", "N")], [("manifest", "CASE_MANIFEST")],
     [("output_path", "path", "", "data/processed/{trade_id}_{asset}_{date}", "Y", "Case folder."),
      ("ready_flag", "bool", "", "true", "Y", "ready_for_extraction.")],
     "agent/email_agent.py (_store_case)", "Built", "Writer / downstream agent")

node("ACT-06", "Report Generator (MIS)", "Action / Output", "Reporting & MIS", "Future",
     [("data", "DATASET", "Y")], [("report", "REPORT")],
     [("format", "enum", "xlsx | pdf | json", "xlsx", "Y", ""),
      ("template", "string", "", "", "N", ""),
      ("kpis", "csv", "", "volume,label_mix,extraction_yield", "N", "")],
     "—", "Capability theme: Reporting & MIS", "Email Sender / Writer")

# —— Integration ——
node("INT-01", "ServiceNow (NEWS) Writer / Closer", "Integration", "Orchestration", "Future",
     [("in", "ANY", "Y")], [("audit", "AUDIT_EVENT")],
     [("instance_url", "url", "", "", "Y", ""),
      ("action", "enum", "create | update | close", "update", "Y", "Trigger/log/close."),
      ("table", "string", "", "incident", "Y", "")],
     "—", "BRD §3 ServiceNow (NEWS) integration", "Audit / end")

# —— Memory / Learning ——
node("MEM-01", "Feedback Loop / Long-term Memory (Vector DB)", "Memory / Learning", "Optimal Refinement", "Future",
     [("decision", "REVIEW_DECISION", "Y")], [("memory", "DATASET")],
     [("store", "enum", "vector-db", "vector-db", "Y", "HITL decisions + reason codes."),
      ("scope", "csv", "", "counterparty,pattern,thresholds", "N", "What memory retains.")],
     "—", "BRD §3 self-learning feedback loop", "Self-Learning Updater / Classifier")

node("MEM-02", "Self-Learning Updater", "Memory / Learning", "Optimal Refinement", "Future",
     [("memory", "REVIEW_DECISION", "Y")], [("config", "DATASET")],
     [("targets", "csv", "", "keywords,thresholds,examples", "Y", "What it recalibrates."),
      ("approval", "bool", "", "true", "N", "Human sign-off before applying.")],
     "—", "BRD: update agent memory over time", "Classifier / Compare & Match")

# —— Governance / Observability ——
node("GOV-01", "Audit Logger", "Governance / Observability", "Orchestration", "Built",
     [("in", "ANY", "Y")], [("event", "AUDIT_EVENT")],
     [("log_dir", "path", "", "logs/audit_<date>.log", "Y", "Append-only."),
      ("ship_to", "enum", "file | SIEM/WORM", "file", "N", "Splunk/Sentinel in prod.")],
     "utils/audit.py", "Built", "any (cross-cutting)")

node("GOV-02", "Tracer / Observability (LangSmith)", "Governance / Observability", "Orchestration", "Future",
     [("in", "ANY", "Y")], [("trace", "AUDIT_EVENT")],
     [("project", "string", "", "", "Y", ""),
      ("sample_rate", "float", "0–1", "1.0", "N", "")],
     "—", "BRD stack: LangSmith", "any (cross-cutting)")

node("GOV-03", "Evaluator (RAGAS / DeepEval)", "Governance / Observability", "Orchestration", "Future",
     [("in", "ANY", "Y")], [("metrics", "REPORT")],
     [("suite", "enum", "RAGAS 2.0 | DeepEval", "RAGAS 2.0", "Y", "Pass gate for promotion."),
      ("thresholds", "json", "", "", "N", "")],
     "—", "BRD stack: RAGAS/DeepEval", "Report / gate")

# —— Utility / Flow ——
node("UTL-01", "Delay / Wait", "Utility / Flow", "Orchestration", "Future",
     [("in", "ANY", "Y")], [("out", "ANY")],
     [("duration", "string", "", "5m", "Y", ""),
      ("until", "string", "", "", "N", "Wait-until timestamp.")],
     "—", "Roadmap", "any")

node("UTL-02", "Merge / Join", "Utility / Flow", "Orchestration", "Future",
     [("in_a", "ANY", "Y"), ("in_b", "ANY", "Y")], [("out", "ANY")],
     [("strategy", "enum", "concat | join-on-key", "concat", "Y", ""),
      ("key", "string", "", "trade_id", "N", "")],
     "—", "Roadmap", "any")

node("UTL-03", "Error Handler / Retry", "Utility / Flow", "Orchestration", "Partial",
     [("in", "ANY", "Y")], [("ok", "ANY"), ("error", "AUDIT_EVENT")],
     [("retries", "int", "", "3", "Y", ""),
      ("on_fail", "enum", "skip | route-error | stop", "route-error", "Y", "Pipeline never breaks on one item.")],
     "per-node try/except (never-raise pattern)", "Partial", "any")


# ───────────────────────── scope: Compare & Match only ─────────────────────────
# The PoC's whole scope IS Compare & Match (both work drivers: FX Alleges & DK and
# SSI Verification). We keep the full end-to-end C&M pipeline and drop only the
# blocks that belong purely to the OTHER capability themes (BRD §3.2 deferred):
#   • Alert / Escalation Sender   → Alerts & Escalation theme
#   • Report Generator (MIS)      → Reporting & MIS theme
CM_DROP = {"ACT-03", "ACT-06"}
N = [n for n in N if n["id"] not in CM_DROP]

# Relabel the "theme" column into clear Compare & Match pipeline stages so the
# whole catalog reads as one coherent C&M capability.
STAGE_BY_CAT = {
    "Source / Trigger":           "Orchestration",
    "Ingest / Parse":             "C&M · Ingest",
    "Classify / Reduce":          "C&M · Classify",
    "Extract / Transform":        "C&M · Extract",
    "Decision / Routing":         "C&M · Route",
    "Compare & Match":            "C&M · Reconcile",
    "Human-in-the-Loop":          "C&M · HITL",
    "Action / Output":            "C&M · Act",
    "Integration":                "C&M · Integrate",
    "Memory / Learning":          "C&M · Feedback",
    "Governance / Observability": "Governance",
    "Utility / Flow":             "Orchestration",
}


def stage(n):
    return STAGE_BY_CAT.get(n["cat"], n["theme"])


# ───────────────────────── workbook helpers ─────────────────────────
wb = Workbook()


def style_header(ws, row, ncols, fill=NAVY):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.alignment = CTR
        cell.border = BORDER


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_row(ws, row, values, wrap=True, fills=None, fonts=None, center_cols=()):
    for c, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.border = BORDER
        cell.alignment = CTR if c in center_cols else (WRAP_TOP if wrap else Alignment(vertical="top"))
        cell.font = Font(size=9)
        if fills and c in fills:
            cell.fill = PatternFill("solid", fgColor=fills[c])
        if fonts and c in fonts:
            cell.font = fonts[c]


def title_block(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(bold=True, size=15, color=WHITE)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    t.fill = PatternFill("solid", fgColor=NAVY_D)
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(row=2, column=2 if False else 1, value=subtitle)
    s.font = Font(italic=True, size=9, color="555555")
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16


def status_cell(value):
    return PatternFill("solid", fgColor=STATUS_FILL.get(value, "FFFFFF"))


# ════════════════════════ Sheet 1: Overview ════════════════════════
ws = wb.active
ws.title = "Overview"
ws.sheet_view.showGridLines = False
set_widths(ws, [3, 30, 70, 18])
title_block(ws, "Workflow Builder — Compare & Match Node Catalog", "Composable agentic workflows for Nomura SSG · the Compare & Match capability (FX Alleges & DK + SSI Verification)", 4)

r = 4
intro = [
    "This workbook is the design catalog for the drag-and-drop workflow builder (the \"agent marketplace\"), scoped to Compare & Match.",
    "Each BLOCK is a reusable node with typed INPUT and OUTPUT ports; you compose workflows by connecting compatible ports.",
    "Compare & Match is ~80% of SSG Ops tasks and the whole PoC scope: ingest → classify → extract → query golden sources → match → score → route → HITL → close → feedback.",
    "Alerts & Escalation, Reporting & MIS and stand-alone Optimal Refinement are OUT of this PoC scope (BRD §3.2) and are excluded.",
]
for line in intro:
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    c = ws.cell(row=r, column=2, value=line)
    c.alignment = LEFT_TOP
    c.font = Font(size=10)
    ws.row_dimensions[r].height = 16
    r += 1

r += 1
ws.cell(row=r, column=2, value="SHEETS").font = Font(bold=True, size=11, color=NAVY)
r += 1
sheet_index = [
    ("Node Catalog", "Master list of every block: category, theme, status, code mapping, typical downstream."),
    ("Node Ports (I O)", "Each node's input and output ports with their data-type."),
    ("Configuration", "Every configurable parameter per node (type, options, default, required)."),
    ("Data Contracts", "The canonical objects that flow on edges (RAW_MESSAGE, TRADE_RECORD, MATCH_RESULT, …)."),
    ("Port Compatibility", "Which output data-type may connect to which input data-type (the connection rules)."),
    ("Connectivity Matrix", "Category-to-category: which block types typically feed which (the permutation grid)."),
    ("Workflow Templates", "Ready-made compositions for the two Compare & Match work drivers."),
    ("Stages & Roadmap", "Compare & Match pipeline stages, build status, and BRD delivery phasing."),
    ("Glossary", "Domain + platform terms."),
]
for name, desc in sheet_index:
    ws.cell(row=r, column=2, value=name).font = Font(bold=True, size=9, color="1F4E79")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws.cell(row=r, column=3, value=desc).font = Font(size=9)
    ws.cell(row=r, column=3).alignment = LEFT_TOP
    r += 1

r += 1
ws.cell(row=r, column=2, value="STATUS LEGEND").font = Font(bold=True, size=11, color=NAVY)
r += 1
legend = [
    ("Built", "Working in the current code (mapped to a module)."),
    ("Partial", "Exists partially (UI-only, or logic without runtime wiring)."),
    ("Planned", "Named/captured in the build; not yet implemented."),
    ("Future", "BRD / roadmap scope beyond the current PoC."),
]
for st, desc in legend:
    cell = ws.cell(row=r, column=2, value=st)
    cell.fill = status_cell(st)
    cell.font = Font(bold=True, size=9, color=STATUS_FONT[st])
    cell.alignment = CTR
    cell.border = BORDER
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws.cell(row=r, column=3, value=desc).font = Font(size=9)
    ws.cell(row=r, column=3).alignment = LEFT_TOP
    r += 1

# counts
r += 1
from collections import Counter
sc = Counter(n["status"] for n in N)
cc = Counter(n["cat"] for n in N)
ws.cell(row=r, column=2, value="AT A GLANCE").font = Font(bold=True, size=11, color=NAVY)
r += 1
ws.cell(row=r, column=2, value=f"Total blocks: {len(N)}").font = Font(size=9, bold=True)
ws.cell(row=r, column=3, value="  ·  ".join(f"{k}: {v}" for k, v in sc.most_common())).font = Font(size=9)
r += 1
ws.cell(row=r, column=2, value=f"Categories: {len(cc)}").font = Font(size=9, bold=True)
ws.cell(row=r, column=3, value=f"Data contracts: {len(CONTRACTS)}").font = Font(size=9)


# ════════════════════════ Sheet 2: Node Catalog ════════════════════════
ws = wb.create_sheet("Node Catalog")
ws.sheet_view.showGridLines = False
cols = ["ID", "Block / Node", "Category", "C&M Stage", "Status",
        "What it does", "Maps to (code) / BRD ref", "Typical downstream"]
widths = [9, 30, 22, 18, 10, 46, 40, 26]
set_widths(ws, widths)
title_block(ws, "Node Catalog — Compare & Match", "Every block in the Compare & Match capability (both work drivers)", len(cols))
hdr = 4
for c, h in enumerate(cols, start=1):
    ws.cell(row=hdr, column=c, value=h)
style_header(ws, hdr, len(cols))
ws.freeze_panes = "A5"

DESC = {
    "TRG-01": "Start a run on demand or via an HTTP webhook call.",
    "TRG-02": "Run the workflow on a fixed cadence (e.g. BRD's 4 cycles / 3 hrs).",
    "TRG-03": "Fire when a ServiceNow (NEWS) case is created/updated.",
    "ING-01": "Pull mailbox messages from Microsoft Graph (OAuth2, paging, attachments).",
    "ING-02": "Read .eml files from a local inbox folder (dev / fast lane).",
    "ING-03": "Poll a mailbox over IMAP.",
    "ING-04": "Call any REST endpoint and return rows.",
    "ING-05": "Read files/attachments from a path or folder.",
    "ING-06": "Ingest SWIFT messages (MT103 / MT202 / MT202COV).",
    "ING-07": "Ingest CSD / ICSD settlement reports.",
    "ING-08": "Ingest a TRAM trade feed.",
    "ING-09": "Run a SQL query against Snowflake.",
    "ING-10": "Parse .xlsx/.csv blotters into one trade row per line (header auto-detect, source-agnostic).",
    "ING-11": "OCR vector/scanned PDFs (e.g. SSI docs) into structured rows.",
    "ING-12": "Extract text/tables from .docx / .txt documents.",
    "CLS-01": "Score & label each email RELEVANT / AMBIGUOUS / IRRELEVANT with transparent keyword rules.",
    "CLS-02": "Classify email intent with an LLM (context-aware, not keyword).",
    "CLS-03": "Skip already-seen messages and duplicate trade IDs (idempotency).",
    "CLS-04": "Map many header spellings to a canonical schema; normalise dates/numbers.",
    "CLS-05": "Explode a blotter into individual trades (or group/sum).",
    "EXT-01": "Pull the trade fields out of a message/attachment (Regex / NLP / AI).",
    "EXT-02": "Turn any unstructured input into structured JSON via an LLM; flag missing fields.",
    "DEC-01": "Branch the flow on a field comparison (true / false).",
    "DEC-02": "Route by confidence (BRD: ≥90% auto-draft, <90% → MO review).",
    "DEC-03": "Route to one of many branches by a field value.",
    "DEC-04": "Compute a per-field-weighted confidence score for a match.",
    "CMP-01": "Look up the trusted record in GLOSS / GLOSS Cash.",
    "CMP-02": "Look up the trusted record in OBI.",
    "CMP-03": "Look up the trusted record in a front-office system.",
    "CMP-04": "Reconcile extracted vs golden trade fields; emit matched / near-match / break + confidence.",
    "CMP-05": "Cross-validate SSIs (sender instruction vs receiver expectation).",
    "CMP-06": "Classify match outcomes into break categories with reasons.",
    "HIL-01": "Pause and route a case to a human; resume on their decision.",
    "HIL-02": "Role-based approve / reject / override gate with full audit.",
    "HIL-03": "Broadcast low-confidence items to Middle Office; capture replies for learning.",
    "ACT-01": "Draft a counterparty response with an LLM (approval-gated).",
    "ACT-02": "Send an email / notification.",
    "ACT-03": "Raise an alert and escalate along a defined path.",
    "ACT-04": "Persist rows to SQLite / Excel / a database.",
    "ACT-05": "Write the per-case folder + manifest.json (Phase-2 handoff).",
    "ACT-06": "Generate an MIS report (volumes, label mix, yield).",
    "INT-01": "Create / update / close a ServiceNow (NEWS) case.",
    "MEM-01": "Store HITL decisions + reason codes in long-term (vector) memory.",
    "MEM-02": "Recalibrate keywords / thresholds / examples from feedback.",
    "GOV-01": "Append a structured compliance audit event (who/what/when/outcome).",
    "GOV-02": "Distributed tracing + cost monitoring (LangSmith).",
    "GOV-03": "Evaluate agent outputs (RAGAS / DeepEval) as a promotion gate.",
    "UTL-01": "Pause for a duration or until a time.",
    "UTL-02": "Merge / join two branches back together.",
    "UTL-03": "Catch errors and retry / route / stop (one bad item never breaks the batch).",
}
row = hdr + 1
for n in sorted(N, key=lambda x: x["id"]):
    ref = n["ref"]
    code_ref = n["code"]
    if ref and ref not in ("Built", "Partial", "Roadmap"):
        code_ref = f"{n['code']}  ·  {ref}"
    vals = [n["id"], n["name"], n["cat"], stage(n), n["status"],
            DESC.get(n["id"], n["name"]), code_ref, n["down"]]
    fills = {3: CAT_FILL.get(n["cat"], "FFFFFF"), 5: STATUS_FILL.get(n["status"], "FFFFFF")}
    fonts = {5: Font(size=9, bold=True, color=STATUS_FONT.get(n["status"], "000000"))}
    write_row(ws, row, vals, fills=fills, fonts=fonts, center_cols=(1, 5))
    row += 1
ws.auto_filter.ref = f"A{hdr}:{get_column_letter(len(cols))}{row-1}"


# ════════════════════════ Sheet 3: Node Ports (I/O) ════════════════════════
ws = wb.create_sheet("Node Ports (I O)")
ws.sheet_view.showGridLines = False
cols = ["ID", "Block / Node", "Status", "Input ports (type)", "Output ports (type)"]
set_widths(ws, [9, 30, 10, 46, 46])
title_block(ws, "Node Ports — Inputs & Outputs", "Typed ports define what can connect to what (see Port Compatibility)", len(cols))
hdr = 4
for c, h in enumerate(cols, start=1):
    ws.cell(row=hdr, column=c, value=h)
style_header(ws, hdr, len(cols))
ws.freeze_panes = "A5"
row = hdr + 1
for n in sorted(N, key=lambda x: x["id"]):
    ins = "\n".join(f"• {p}  →  [{t}]{'  (required)' if req=='Y' else ''}" for p, t, req in n["ins"]) or "— (trigger / entry node)"
    outs = "\n".join(f"• {p}  →  [{t}]" for p, t in n["outs"]) or "—"
    fills = {3: STATUS_FILL.get(n["status"], "FFFFFF")}
    fonts = {3: Font(size=9, bold=True, color=STATUS_FONT.get(n["status"], "000000"))}
    write_row(ws, row, [n["id"], n["name"], n["status"], ins, outs], fills=fills, fonts=fonts, center_cols=(1, 3))
    row += 1
ws.auto_filter.ref = f"A{hdr}:{get_column_letter(len(cols))}{row-1}"


# ════════════════════════ Sheet 4: Configuration ════════════════════════
ws = wb.create_sheet("Configuration")
ws.sheet_view.showGridLines = False
cols = ["Node ID", "Block / Node", "Parameter", "Type", "Options / Allowed", "Default", "Req", "Notes"]
set_widths(ws, [9, 26, 22, 10, 28, 26, 6, 40])
title_block(ws, "Node Configuration", "The side-panel parameters for every block", len(cols))
hdr = 4
for c, h in enumerate(cols, start=1):
    ws.cell(row=hdr, column=c, value=h)
style_header(ws, hdr, len(cols))
ws.freeze_panes = "A5"
row = hdr + 1
for n in sorted(N, key=lambda x: x["id"]):
    first = True
    for (param, typ, opts, dflt, req, notes) in n["cfg"]:
        idc = n["id"] if first else ""
        namec = n["name"] if first else ""
        write_row(ws, row, [idc, namec, param, typ, opts, dflt, req, notes], center_cols=(1, 4, 7))
        if first:
            ws.cell(row=row, column=1).font = Font(size=9, bold=True)
        first = False
        row += 1
ws.auto_filter.ref = f"A{hdr}:{get_column_letter(len(cols))}{row-1}"


# ════════════════════════ Sheet 5: Data Contracts ════════════════════════
ws = wb.create_sheet("Data Contracts")
ws.sheet_view.showGridLines = False
cols = ["Data type (port)", "What it is", "Key fields", "Produced by", "Consumed by"]
set_widths(ws, [20, 40, 50, 26, 30])
title_block(ws, "Data Contracts", "The payload objects that travel along the edges between nodes", len(cols))
hdr = 4
for c, h in enumerate(cols, start=1):
    ws.cell(row=hdr, column=c, value=h)
style_header(ws, hdr, len(cols))
ws.freeze_panes = "A5"
row = hdr + 1
for (t, what, fields, prod, cons) in CONTRACTS:
    write_row(ws, row, [t, what, fields, prod, cons], center_cols=(1,))
    ws.cell(row=row, column=1).font = Font(size=9, bold=True, color=NAVY)
    row += 1
ws.auto_filter.ref = f"A{hdr}:{get_column_letter(len(cols))}{row-1}"


# ════════════════════════ Sheet 6: Port Compatibility ════════════════════════
ws = wb.create_sheet("Port Compatibility")
ws.sheet_view.showGridLines = False
title_block(ws, "Port Compatibility Matrix", "Rows = OUTPUT type · Columns = INPUT type · ●=exact  ◐=adapt/cast  ＊=ANY wildcard", len(CONTRACT_TYPES) + 1)
hdr = 4
ws.cell(row=hdr, column=1, value="OUT ▼  /  IN ▶")
for j, t in enumerate(CONTRACT_TYPES, start=2):
    ws.cell(row=hdr, column=j, value=t)
style_header(ws, hdr, len(CONTRACT_TYPES) + 1)
set_widths(ws, [16] + [13] * len(CONTRACT_TYPES))
ws.freeze_panes = "B5"
row = hdr + 1
for ot in CONTRACT_TYPES:
    ws.cell(row=row, column=1, value=ot).font = Font(bold=True, size=9, color=WHITE)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=GREY_HDR)
    ws.cell(row=row, column=1).alignment = CTR
    ws.cell(row=row, column=1).border = BORDER
    for j, it in enumerate(CONTRACT_TYPES, start=2):
        kind = compatible(ot, it)
        mark = {"EXACT": "●", "ADAPT": "◐", "ANY": "＊"}.get(kind, "")
        cell = ws.cell(row=row, column=j, value=mark)
        cell.alignment = CTR
        cell.border = BORDER
        cell.font = Font(size=11, bold=True,
                         color={"EXACT": "1E6B34", "ADAPT": "9C6500", "ANY": "1F4E79"}.get(kind, "999999"))
        if kind == "EXACT":
            cell.fill = PatternFill("solid", fgColor="E2EFDA")
        elif kind == "ADAPT":
            cell.fill = PatternFill("solid", fgColor="FFF2CC")
        elif kind == "ANY":
            cell.fill = PatternFill("solid", fgColor="DDEBF7")
    row += 1


# ════════════════════════ Sheet 7: Connectivity Matrix (category) ════════════════════════
ws = wb.create_sheet("Connectivity Matrix")
ws.sheet_view.showGridLines = False
CATS = list(CAT_FILL.keys())
# typical category → category flow
FLOW = {
    "Source / Trigger":           ["Ingest / Parse"],
    "Ingest / Parse":             ["Classify / Reduce", "Extract / Transform", "Utility / Flow", "Action / Output"],
    "Classify / Reduce":          ["Decision / Routing", "Extract / Transform", "Compare & Match", "Human-in-the-Loop", "Action / Output"],
    "Extract / Transform":        ["Compare & Match", "Decision / Routing", "Action / Output", "Classify / Reduce"],
    "Decision / Routing":         ["Extract / Transform", "Compare & Match", "Human-in-the-Loop", "Action / Output", "Integration"],
    "Compare & Match":            ["Decision / Routing", "Human-in-the-Loop", "Action / Output", "Memory / Learning"],
    "Human-in-the-Loop":          ["Memory / Learning", "Action / Output", "Integration", "Compare & Match"],
    "Action / Output":            ["Integration", "Governance / Observability"],
    "Integration":                ["Governance / Observability"],
    "Memory / Learning":          ["Classify / Reduce", "Compare & Match"],
    "Governance / Observability": [],
    "Utility / Flow":             CATS[:],   # plumbing connects to anything
}
title_block(ws, "Connectivity Matrix (category)", "Rows feed columns · ✔=typical flow · ○=possible · blank=not typical · Governance/Utility are cross-cutting", len(CATS) + 1)
hdr = 4
ws.cell(row=hdr, column=1, value="FROM ▼ / TO ▶")
for j, c in enumerate(CATS, start=2):
    ws.cell(row=hdr, column=j, value=c)
style_header(ws, hdr, len(CATS) + 1)
set_widths(ws, [24] + [11] * len(CATS))
ws.row_dimensions[hdr].height = 54
for j in range(2, len(CATS) + 2):
    ws.cell(row=hdr, column=j).alignment = Alignment(text_rotation=90, horizontal="center", vertical="bottom", wrap_text=True)
ws.freeze_panes = "B5"
row = hdr + 1
for fr in CATS:
    ws.cell(row=row, column=1, value=fr).font = Font(bold=True, size=8, color=WHITE)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=GREY_HDR)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.cell(row=row, column=1).border = BORDER
    typ = set(FLOW.get(fr, []))
    for j, to in enumerate(CATS, start=2):
        mark = ""
        if to in typ:
            mark = "✔"
        elif fr == "Utility / Flow" or to in ("Governance / Observability",):
            mark = "○"
        cell = ws.cell(row=row, column=j, value=mark)
        cell.alignment = CTR
        cell.border = BORDER
        if mark == "✔":
            cell.fill = PatternFill("solid", fgColor="C6EFCE")
            cell.font = Font(bold=True, size=10, color="1E6B34")
        elif mark == "○":
            cell.fill = PatternFill("solid", fgColor="F2F2F2")
            cell.font = Font(size=10, color="888888")
    row += 1


# ════════════════════════ Sheet 8: Workflow Templates ════════════════════════
ws = wb.create_sheet("Workflow Templates")
ws.sheet_view.showGridLines = False
cols = ["Template / use-case", "Work driver / Theme", "Step", "Block", "Key config", "→ produces", "Status"]
set_widths(ws, [26, 22, 6, 30, 34, 22, 10])
title_block(ws, "Workflow Templates — Compare & Match", "Ready-made block compositions (real permutations) for the two Compare & Match work drivers", len(cols))
hdr = 4
for c, h in enumerate(cols, start=1):
    ws.cell(row=hdr, column=c, value=h)
style_header(ws, hdr, len(cols))
ws.freeze_panes = "A5"

TEMPLATES = [
    ("Trade Extraction Pipeline (current PoC)", "FX Settlement", "Built", [
        ("Email Ingestor — MS Graph API", "source=Graph, store=FS", "RAW_MESSAGE"),
        ("Data Reducer / Classifier (NLP)", "mode=NLP, thr=0.7/0.3", "CLASSIFIED_ITEM"),
        ("Decider", "confidence_score > 80", "branch"),
        ("Extractor (true)", "method=Regex, 6 fields", "TRADE_RECORD"),
        ("Man in the Loop (false)", "email ops, pause", "REVIEW_DECISION"),
        ("Case / Manifest Builder", "ready_for_extraction", "CASE_MANIFEST"),
    ]),
    ("FX Alleges & DK", "Work driver 1", "Future", [
        ("Email Ingestor — MS Graph API", "$filter Allege/DK", "RAW_MESSAGE"),
        ("Intent Classifier (AI)", "labels=Allege,DK,…", "CLASSIFIED_ITEM"),
        ("Unstructured→Structured Extractor", "schema=6 fields", "TRADE_RECORD"),
        ("Golden Source Query — GLOSS", "match trade_id", "GOLDEN_RECORD"),
        ("Compare & Match Engine", "5 fields, fuzzy", "MATCH_RESULT"),
        ("Confidence Router", "≥90 auto / <90 MO", "branch"),
        ("Response Drafter (≥90)", "auto-draft reply", "NOTIFICATION"),
        ("EMODEST Broadcast (<90)", "MO feedback loop", "REVIEW_DECISION"),
        ("ServiceNow (NEWS) Writer", "log + close", "AUDIT_EVENT"),
    ]),
    ("SSI Verification", "Work driver 2", "Future", [
        ("File / Folder Ingestor", "SSI PDFs", "DOCUMENT"),
        ("PDF / OCR Extractor", "ocr=true", "TRADE_RECORD"),
        ("SSI Cross-Validator", "vs GLOSS Cash", "MATCH_RESULT"),
        ("Exception / Break Classifier", "flag mismatches", "MATCH_RESULT"),
        ("Decider", "break == true", "branch"),
        ("Approval Gate (RBAC)", "reviewer approves", "REVIEW_DECISION"),
        ("ServiceNow (NEWS) Writer", "confirm/close", "AUDIT_EVENT"),
    ]),
    ("Self-learning feedback loop", "C&M · Feedback", "Future", [
        ("Man in the Loop / Approval Gate", "human corrections", "REVIEW_DECISION"),
        ("Feedback Loop / Memory", "store decisions + reasons", "DATASET"),
        ("Self-Learning Updater", "recalibrate thresholds / match rules", "DATASET"),
        ("Compare & Match Engine", "applies sharpened config next run", "MATCH_RESULT"),
    ]),
]
row = hdr + 1
band = False
for name, driver, status, steps in TEMPLATES:
    band = not band
    start = row
    for i, (block, cfg, out) in enumerate(steps, start=1):
        namec = name if i == 1 else ""
        drvc = driver if i == 1 else ""
        stc = status if i == 1 else ""
        write_row(ws, row, [namec, drvc, i, block, cfg, out, stc], center_cols=(3, 7))
        if band:
            for cc in range(1, len(cols) + 1):
                if ws.cell(row=row, column=cc).fill.fgColor.rgb in (None, "00000000"):
                    ws.cell(row=row, column=cc).fill = PatternFill("solid", fgColor="F7F9FC")
        if i == 1:
            ws.cell(row=row, column=1).font = Font(size=9, bold=True, color=NAVY)
            ws.cell(row=row, column=7).fill = status_cell(status)
            ws.cell(row=row, column=7).font = Font(size=9, bold=True, color=STATUS_FONT.get(status, "000000"))
        row += 1
    # merge the name/driver/status cells for the group
    ws.merge_cells(start_row=start, start_column=1, end_row=row - 1, end_column=1)
    ws.merge_cells(start_row=start, start_column=2, end_row=row - 1, end_column=2)
    ws.merge_cells(start_row=start, start_column=7, end_row=row - 1, end_column=7)
    ws.cell(row=start, column=1).alignment = Alignment(vertical="center", wrap_text=True)
    ws.cell(row=start, column=2).alignment = Alignment(vertical="center", wrap_text=True)
    ws.cell(row=start, column=7).alignment = CTR


# ════════════════════════ Sheet 9: Stages & Roadmap ════════════════════════
ws = wb.create_sheet("Stages & Roadmap")
ws.sheet_view.showGridLines = False
cols = ["C&M stage", "What happens", "Representative blocks", "Build status"]
set_widths(ws, [20, 48, 46, 16])
title_block(ws, "Compare & Match — Stages & Roadmap", "The end-to-end C&M pipeline (≈80% of SSG Ops tasks) broken into stages", len(cols))
hdr = 4
for c, h in enumerate(cols, start=1):
    ws.cell(row=hdr, column=c, value=h)
style_header(ws, hdr, len(cols))
ws.freeze_panes = "A5"
stages_tbl = [
    ("Ingest", "Pull source data: counterparty emails, SWIFT, CSD/ICSD, TRAM, files, Snowflake/REST.",
     "Email/IMAP/SWIFT/CSD-ICSD/TRAM/File/Snowflake Ingestors, Attachment Parser, PDF-OCR",
     "Built→Future"),
    ("Classify", "Identify intent (Allege / DK / SSI / noise); dedupe; normalise to a canonical schema.",
     "Classifier (NLP), Intent Classifier (AI), Deduplicator, Normalizer, Aggregator",
     "Built→Planned"),
    ("Extract", "Convert unstructured input into the 6 structured trade fields (JSON).",
     "Field Extractor (Regex/NLP/AI), Unstructured→Structured Extractor (LLM)",
     "Partial→Future"),
    ("Reconcile", "Query golden sources and compare field-by-field with a confidence score.",
     "Golden Source Query (GLOSS/OBI/FO), Compare & Match Engine, SSI Cross-Validator, Break Classifier, Confidence Scorer",
     "Future"),
    ("Route", "Branch on confidence: ≥90% auto-draft, <90% to Middle Office.",
     "Decider, Confidence Router, Switch",
     "Partial→Future"),
    ("HITL", "Mandatory human review/approve/override gate before any outbound action.",
     "Man in the Loop, Approval Gate (RBAC), EMODEST Broadcast",
     "Partial→Future"),
    ("Act / Close", "Draft the counterparty response, persist the case, log & close in ServiceNow.",
     "Response Drafter, Email Sender, Data Writer, Case/Manifest Builder, ServiceNow Writer",
     "Built→Future"),
    ("Feedback", "Capture HITL decisions in memory and recalibrate the agents over time.",
     "Feedback Loop / Memory, Self-Learning Updater",
     "Future"),
    ("Govern", "Audit trail, tracing and evaluation gates across every stage.",
     "Audit Logger, Tracer (LangSmith), Evaluator (RAGAS/DeepEval)",
     "Built→Future"),
]
row = hdr + 1
for t, cov, blocks, st in stages_tbl:
    write_row(ws, row, [t, cov, blocks, st], center_cols=(4,))
    ws.cell(row=row, column=1).font = Font(size=10, bold=True, color=NAVY)
    row += 1

row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
c = ws.cell(row=row, column=1, value="Out of this PoC's scope (BRD §3.2, separate themes): Alerts & Escalation · Reporting & MIS · stand-alone Optimal Refinement.")
c.font = Font(italic=True, size=9, color="9C6500")
c.fill = PatternFill("solid", fgColor="FFF7E6")
c.alignment = LEFT_TOP
c.border = BORDER
row += 2

ws.cell(row=row, column=1, value="DELIVERY PHASING (BRD 5 modules + roadmap)").font = Font(bold=True, size=11, color=NAVY)
row += 1
phases = [
    ("Module 1 · Discovery & Setup (wk 1-2)", "Walkthroughs, golden sources, SOPs, RBAC, synthetic data, env access."),
    ("Module 2 · Agent Build & Iterative Testing (wk 3-7)", "Build extraction / compare & match / orchestration / HITL / feedback + UI; marketplace scaffolding."),
    ("Module 3 · Integration & E2E (wk 8-10)", "ServiceNow trigger/log/close; MO mailbox feedback loop; E2E on real data."),
    ("Module 4 · UAT & Go-Live (wk 10-11)", "UAT on Nomura env; defect fixes; go-live on sign-off."),
    ("Module 5 · Knowledge Transfer (wk 12)", "Architecture, agent design, runbooks, repo handover."),
    ("Builder / Marketplace (post-PoC)", "This catalog → drag-drop canvas; reusable capability blocks; per-domain config."),
]
for p, d in phases:
    ws.cell(row=row, column=1, value=p).font = Font(size=9, bold=True, color="1F4E79")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws.cell(row=row, column=2, value=d).font = Font(size=9)
    ws.cell(row=row, column=2).alignment = LEFT_TOP
    row += 1


# ════════════════════════ Sheet 10: Glossary ════════════════════════
ws = wb.create_sheet("Glossary")
ws.sheet_view.showGridLines = False
cols = ["Term", "Meaning"]
set_widths(ws, [26, 90])
title_block(ws, "Glossary", "Domain + platform terms", len(cols))
hdr = 4
for c, h in enumerate(cols, start=1):
    ws.cell(row=hdr, column=c, value=h)
style_header(ws, hdr, len(cols))
ws.freeze_panes = "A5"
glossary = [
    ("Alleges & DK", "Counterparty asserts a trade (Allege); receiver disputes/doesn't recognise it (Don't Know)."),
    ("SSI", "Standard Settlement Instructions — where/how to settle a trade (often in PDFs)."),
    ("SSI Verification", "Retrieve SSIs, cross-validate sender vs receiver expectation (e.g. GLOSS Cash), flag exceptions."),
    ("Compare & Match", "Reconcile extracted trade fields against golden sources with a confidence score (~80% of SSG tasks)."),
    ("Golden source", "Trusted system of record: GLOSS API / GLOSS Cash, OBI, FO systems."),
    ("EMODEST broadcast", "MO broadcast for <90% confidence items; replies feed the learning loop via coded subjects."),
    ("HITL", "Human-in-the-Loop — mandatory review/approve/override gate before any outbound action."),
    ("RBAC", "Role-Based Access Control over who can view/approve/override/retire agent actions."),
    ("SPAR", "Sense → Plan → Act → Reflect — the agent loop; here realised as the workflow graph."),
    ("Node / Block", "A reusable capability unit with typed input/output ports and a config panel."),
    ("Port", "A typed connection point on a node; outputs connect to compatible inputs."),
    ("Data contract", "The shape of the payload travelling on an edge (RAW_MESSAGE, TRADE_RECORD, …)."),
    ("Agent marketplace", "Library of reusable capability blocks composed into workflows per use-case."),
    ("TRAM / SWIFT", "Trade feed / interbank messaging (MT103, MT202, MT202COV) — ingestion sources."),
    ("CSD / ICSD", "(International) Central Securities Depository reports — settlement source data."),
    ("Chinou", "Nomura's enterprise Anthropic (Claude) instance used for AI nodes."),
    ("ServiceNow (NEWS)", "Nomura's workflow/exception platform; triggers, logging, and case closure."),
    ("RAGAS / DeepEval", "Agent-output evaluation frameworks; a pass gate before production promotion."),
    ("Manifest", "Per-case JSON handoff contract (trade_id, paths, classification, extraction, ready flag)."),
    ("Idempotent", "Re-running yields the same state; guarded by message_id + trade_id dedup."),
]
row = hdr + 1
for term, mean in glossary:
    write_row(ws, row, [term, mean])
    ws.cell(row=row, column=1).font = Font(size=9, bold=True, color=NAVY)
    row += 1
ws.auto_filter.ref = f"A{hdr}:B{row-1}"


# ───────────────────────── save ─────────────────────────
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"Wrote {OUT}  ({len(N)} nodes, {len(wb.sheetnames)} sheets)")
print("Sheets:", ", ".join(wb.sheetnames))
